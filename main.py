import os
import time
import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.image import Image
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.popup import Popup
from kivy.core.window import Window
from kivy.metrics import dp
from kivy.utils import platform
from kivy.clock import Clock
from kivy.graphics import Color, Rectangle

# --- VISUAL THEME ---
COLOR_BG = (0.1, 0.1, 0.1, 1)       
COLOR_ACCENT = (0, 0.8, 0.4, 1)     
COLOR_TEXT = (1, 1, 1, 1)           
COLOR_SUBTEXT = (0.7, 0.7, 0.7, 1)  
COLOR_CARD_BG = (0.15, 0.15, 0.15, 1)

# --- CONFIGURATION ---
GLOBAL_SETTINGS = {
    "exchange_rate": 36.0,    
    "shipping_rate": 50000.0, 
    "margin_percent": 0.0     
}

SESSION_STATE = {
    "data": [],
    "filepath": None,
    "header_row": 1,
    "col_map": {},
    "total_investment": 0
}

# --- PERMISSION LOGIC ---
def request_android_permissions():
    """ Only runs on Android. Safe to ignore on PC. """
    if platform == 'android':
        try:
            from jnius import autoclass
            from android.permissions import request_permissions, Permission
            Build = autoclass('android.os.Build')
            VERSION = autoclass('android.os.Build$VERSION')
            if VERSION.SDK_INT >= 30:
                Environment = autoclass('android.os.Environment')
                if not Environment.isExternalStorageManager():
                    Intent = autoclass('android.content.Intent')
                    Settings = autoclass('android.provider.Settings')
                    Uri = autoclass('android.net.Uri')
                    PythonActivity = autoclass('org.kivy.android.PythonActivity')
                    intent = Intent(Settings.ACTION_MANAGE_APP_ALL_FILES_ACCESS_PERMISSION)
                    activity = PythonActivity.mActivity
                    package_uri = Uri.parse("package:" + activity.getPackageName())
                    intent.setData(package_uri)
                    activity.startActivity(intent)
            else:
                request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])
        except Exception as e:
            print(f"Permission Error: {e}")

# --- LOGIC ENGINE ---
def process_excel_preserve_images(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        
        # 1. EXTRACT IMAGES
        image_map = {}
        temp_dir = os.path.abspath(App.get_running_app().user_data_dir)
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        
        if hasattr(sheet, 'images'):
            for idx, img in enumerate(sheet.images):
                try:
                    # Robust Anchor Logic
                    if hasattr(img.anchor, '_from'):
                        row = img.anchor._from.row + 1
                    elif hasattr(img.anchor, 'row'):
                        row = img.anchor.row + 1
                    else:
                        continue 

                    img_name = f"img_{row}_{idx}.png"
                    img_path = os.path.join(temp_dir, img_name)
                    
                    if hasattr(img, 'ref'): # Pillow
                        img.ref.save(img_path)
                    else:
                        img.save(img_path)
                        
                    image_map[row] = img_path
                except Exception as e:
                    print(f"Img Error: {e}")

        # 2. FIND HEADER
        header_row_index = -1
        col_map = {}
        for r in range(1, 20):
            row_values = [str(cell.value).strip() if cell.value else "" for cell in sheet[r]]
            if "ITEM" in row_values or "Price(RMB)" in row_values:
                header_row_index = r
                for idx, val in enumerate(row_values):
                    col_map[val] = idx + 1 
                break
        
        if header_row_index == -1: return None, "Header not found."

        SESSION_STATE["filepath"] = filepath
        SESSION_STATE["header_row"] = header_row_index
        SESSION_STATE["col_map"] = col_map

        results = []
        grand_total_dzd = 0
        
        # 3. PROCESS ROWS
        for r in range(header_row_index + 1, sheet.max_row + 1):
            try:
                def get_val(name):
                    return sheet.cell(row=r, column=col_map[name]).value if name in col_map else 0

                item_name = str(sheet.cell(row=r, column=col_map.get("ITEM", 1)).value or "")
                if item_name == "None" or item_name == "":
                     item_name = str(sheet.cell(row=r, column=col_map.get("المنتوج", 1)).value or "Unknown")

                rmb_price = float(get_val("Price(RMB)") or 0)
                boxes_count = float(get_val("Ctn") or 0)
                units_per_box = float(get_val("Qty") or 1) 
                cbm_per_box = float(get_val("CBM") or 0)

                if boxes_count == 0: continue

                # LOGIC: (CBM * Rate) / Qty
                shipping_cost_box = cbm_per_box * GLOBAL_SETTINGS["shipping_rate"]
                shipping_per_unit = shipping_cost_box / units_per_box
                
                product_base_dzd = rmb_price * GLOBAL_SETTINGS["exchange_rate"]
                final_unit_cost = product_base_dzd + shipping_per_unit
                
                total_line_cost = final_unit_cost * (boxes_count * units_per_box)
                grand_total_dzd += total_line_cost

                img_path = image_map.get(r, None)

                results.append({
                    "row_index": r,
                    "name": item_name,
                    "unit_cost": round(final_unit_cost, 2),
                    "total_line": round(total_line_cost, 2),
                    "rmb_price": rmb_price,
                    "qty": int(boxes_count * units_per_box),
                    "image": img_path 
                })

            except Exception:
                continue
        
        SESSION_STATE["total_investment"] = grand_total_dzd
        return results, "Success"

    except Exception as e:
        return None, str(e)

def export_results_smart():
    data = SESSION_STATE["data"]
    filepath = SESSION_STATE["filepath"]
    
    if not data or not filepath: return False, "No data"

    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        bold_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        h_row = SESSION_STATE["header_row"]

        # 1. IDENTIFY COLUMNS
        ctn_col = None
        total_col = None
        
        # Scan headers to find Ctn and Total
        for cell in sheet[h_row]:
            val = str(cell.value).strip().lower()
            if val in ["ctn", "carton", "box", "boxes", "qty (ctn)"]:
                ctn_col = cell.column
            elif "total" in val or "amount" in val:
                total_col = cell.column
        
        # Fallback if Total not found -> Use last column + 1
        if not total_col:
            total_col = sheet.max_column + 1

        # Fallback if Ctn not found -> Look up in col_map
        if not ctn_col:
            ctn_col = SESSION_STATE["col_map"].get("Ctn")

        if ctn_col:
            # 2. MOVE CTN DATA TO TOTAL COLUMN
            # Header
            cell_total_header = sheet.cell(row=h_row, column=total_col)
            cell_total_header.value = "Ctn" # Rename Total to Ctn
            cell_total_header.font = bold_font
            cell_total_header.alignment = Alignment(horizontal="center")

            # Data Move Loop
            # We move the *original* values from Ctn column to Total column
            for row in range(h_row + 1, sheet.max_row + 1):
                old_val = sheet.cell(row=row, column=ctn_col).value
                sheet.cell(row=row, column=total_col).value = old_val
                # Add basic border
                sheet.cell(row=row, column=total_col).border = thin_border
                sheet.cell(row=row, column=total_col).alignment = Alignment(horizontal="center")

            # 3. OVERWRITE CTN COLUMN WITH PRICE
            # Header
            cell_price_header = sheet.cell(row=h_row, column=ctn_col)
            cell_price_header.value = "Unit Cost (DZD)"
            cell_price_header.font = bold_font
            cell_price_header.fill = fill # Blue BG
            cell_price_header.alignment = Alignment(horizontal="center")

            # Fill Prices
            for item in data:
                r = item["row_index"]
                cell = sheet.cell(row=r, column=ctn_col)
                cell.value = item["unit_cost"]
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        else:
            # Emergency Fallback if no Ctn column exists at all (Insert after name)
            # This shouldn't happen with your file structure
            name_col = SESSION_STATE["col_map"].get("ITEM", 2)
            target_col = name_col + 1
            sheet.insert_cols(target_col)
            sheet.cell(row=h_row, column=target_col).value = "Unit Cost (DZD)"
            for item in data:
                sheet.cell(row=item["row_index"], column=target_col).value = item["unit_cost"]

        timestamp = int(time.time())
        if platform == 'android':
            output_name = f"/storage/emulated/0/Download/CostSheet_{timestamp}.xlsx"
        else:
            output_name = os.path.join(os.path.expanduser("~"), "Downloads", f"CostSheet_{timestamp}.xlsx")

        wb.save(output_name)
        return True, output_name

    except Exception as e:
        return False, str(e)

# --- UI COMPONENTS ---
class GalleryCard(BoxLayout):
    def __init__(self, item, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.size_hint_y = None
        self.height = dp(220) 
        self.padding = 5
        self.spacing = 5
        with self.canvas.before:
            Color(*COLOR_CARD_BG)
            self.rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self.update_rect, size=self.update_rect)

        img_source = item.get('image')
        if img_source:
            img = Image(source=img_source, size_hint_y=0.6, allow_stretch=True, keep_ratio=True)
            self.add_widget(img)
        else:
            self.add_widget(Label(text="No Image", size_hint_y=0.6, color=(0.5,0.5,0.5,1)))

        self.add_widget(Label(text=item['name'][:15] + "...", size_hint_y=0.2, color=COLOR_TEXT, bold=True))
        self.add_widget(Label(text=f"{item['unit_cost']} DA", size_hint_y=0.2, color=COLOR_ACCENT, bold=True, font_size='18sp'))

    def update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size

class InfoCard(BoxLayout):
    def __init__(self, item, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'horizontal' 
        self.size_hint_y = None
        self.height = dp(120)
        self.padding = 10
        self.spacing = 10
        
        img_source = item.get('image')
        if img_source:
            img = Image(source=img_source, size_hint_x=0.3, allow_stretch=True, keep_ratio=True)
            self.add_widget(img)
        else:
            lbl = Label(text="No IMG", size_hint_x=0.3, color=(0.5,0.5,0.5,1))
            self.add_widget(lbl)

        text_box = BoxLayout(orientation='vertical')
        top = BoxLayout()
        lbl_name = Label(text=f"[b]{item['name']}[/b]", markup=True, halign="left", valign="middle", color=COLOR_TEXT, font_size='16sp')
        lbl_name.bind(size=lbl_name.setter('text_size'))
        lbl_cost = Label(text=f"{item['unit_cost']} DA", color=COLOR_ACCENT, font_size='20sp', bold=True, size_hint_x=0.4, halign='right')
        top.add_widget(lbl_name)
        top.add_widget(lbl_cost)
        
        bot = BoxLayout()
        lbl_detail = Label(text=f"Qty: {item['qty']} | RMB: {item['rmb_price']}", color=COLOR_SUBTEXT, font_size='13sp')
        lbl_total = Label(text=f"Total: {int(item['total_line']):,} DA", color=COLOR_SUBTEXT, font_size='13sp', halign='right')
        bot.add_widget(lbl_detail)
        bot.add_widget(lbl_total)
        
        text_box.add_widget(top)
        text_box.add_widget(bot)
        self.add_widget(text_box)

class TableRow(BoxLayout):
    def __init__(self, item, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'horizontal'
        self.size_hint_y = None
        self.height = dp(40)
        self.spacing = 10
        self.add_widget(Label(text=item['name'][:20], size_hint_x=0.5, halign='left', color=COLOR_TEXT))
        self.add_widget(Label(text=str(item['unit_cost']), size_hint_x=0.2, color=COLOR_ACCENT, bold=True))
        self.add_widget(Label(text=str(item['qty']), size_hint_x=0.15, color=COLOR_SUBTEXT))
        self.add_widget(Label(text=f"{int(item['total_line']):,}", size_hint_x=0.25, color=COLOR_SUBTEXT))

# --- SCREENS ---

class HomeScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=40, spacing=30)
        layout.add_widget(Label(text="IMPORT CALCULATOR", font_size='28sp', bold=True, color=COLOR_ACCENT, size_hint=(1, 0.2)))
        layout.add_widget(Label(text="DZD Groupage Edition", font_size='16sp', color=COLOR_SUBTEXT, size_hint=(1, 0.1)))

        # 1. PRICE CHECKER (UPDATED)
        btn_conv = Button(text="🧮 QUICK PRICE CHECKER", size_hint=(1, 0.15), background_color=(0.1, 0.3, 0.5, 1), font_size='16sp')
        btn_conv.bind(on_press=self.open_converter)
        layout.add_widget(btn_conv)

        btn_import = Button(text="📂 IMPORT EXCEL", size_hint=(1, 0.2), background_color=(0.2, 0.2, 0.2, 1), font_size='18sp')
        btn_import.bind(on_press=self.show_file_chooser)
        layout.add_widget(btn_import)

        btn_settings = Button(text="⚙️ SETTINGS", size_hint=(1, 0.15), background_color=(0.1, 0.1, 0.1, 1), color=COLOR_SUBTEXT)
        btn_settings.bind(on_press=self.go_settings)
        layout.add_widget(btn_settings)
        self.add_widget(layout)

    def go_settings(self, instance): self.manager.current = 'settings'

    # --- UPDATED CONVERTER: NOW WITH QUANTITY ---
    def open_converter(self, instance):
        content = BoxLayout(orientation='vertical', padding=20, spacing=15)
        
        # Row 1: RMB
        row1 = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        row1.add_widget(Label(text="RMB (Unit):", size_hint_x=0.4, color=COLOR_TEXT))
        inp_rmb = TextInput(multiline=False, input_filter='float', hint_text="Price", write_tab=False)
        row1.add_widget(inp_rmb)
        
        # Row 2: CBM
        row2 = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        row2.add_widget(Label(text="CBM (Box):", size_hint_x=0.4, color=COLOR_TEXT))
        inp_cbm = TextInput(multiline=False, input_filter='float', hint_text="Total Volume", text="0", write_tab=False)
        row2.add_widget(inp_cbm)
        
        # Row 3: Qty (THE NEW FEATURE)
        row3 = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        row3.add_widget(Label(text="Qty (Box):", size_hint_x=0.4, color=COLOR_TEXT))
        inp_qty = TextInput(multiline=False, input_filter='int', hint_text="Pieces per Box", text="1", write_tab=False)
        row3.add_widget(inp_qty)
        
        lbl_result = Label(text="0.00 DZD", font_size='28sp', color=COLOR_ACCENT, bold=True, size_hint_y=0.3)
        
        def calculate(inst):
            try:
                rmb = float(inp_rmb.text) if inp_rmb.text else 0
                cbm = float(inp_cbm.text) if inp_cbm.text else 0
                qty = float(inp_qty.text) if inp_qty.text else 1
                if qty == 0: qty = 1
                
                # Formula: (CBM * 50000 / Qty) + (RMB * 36)
                shipping_box = cbm * GLOBAL_SETTINGS["shipping_rate"]
                shipping_unit = shipping_box / qty
                base_cost = rmb * GLOBAL_SETTINGS["exchange_rate"]
                
                total = base_cost + shipping_unit
                lbl_result.text = f"{total:,.2f} DA"
            except:
                lbl_result.text = "Error"

        btn_calc = Button(text="CALCULATE COST", background_color=COLOR_ACCENT, size_hint_y=None, height=dp(50), bold=True)
        btn_calc.bind(on_press=calculate)
        
        content.add_widget(row1)
        content.add_widget(row2)
        content.add_widget(row3)
        content.add_widget(btn_calc)
        content.add_widget(lbl_result)
        
        popup = Popup(title="Landing Cost Calculator", content=content, size_hint=(0.9, 0.6))
        popup.open()
    
    def show_file_chooser(self, instance):
        request_android_permissions()
        content = BoxLayout(orientation='vertical')
        
        if platform == 'android':
            start_path = '/storage/emulated/0'
        else:
            start_path = os.path.expanduser("~")
            
        filechooser = FileChooserIconView(path=start_path, filters=['*.xlsx'])
        btn_box = BoxLayout(size_hint_y=0.1)
        btn_load = Button(text="Load")
        btn_cancel = Button(text="Cancel")
        content.add_widget(filechooser)
        content.add_widget(btn_box)
        btn_box.add_widget(btn_cancel)
        btn_box.add_widget(btn_load)
        popup = Popup(title="Select File", content=content, size_hint=(0.9, 0.9))
        
        def load(inst):
            if filechooser.selection:
                data, status = process_excel_preserve_images(filechooser.selection[0])
                if data:
                    SESSION_STATE["data"] = data
                    self.manager.get_screen('results').load_data()
                    self.manager.current = 'results'
                    popup.dismiss()
                else: print(status)
        btn_load.bind(on_press=load)
        btn_cancel.bind(on_press=popup.dismiss)
        popup.open()

class ResultsScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        self.view_mode = "list" 
        
        dash = BoxLayout(size_hint_y=None, height=dp(80), padding=10, spacing=10)
        btn_back = Button(text="<", size_hint_x=None, width=dp(50), background_color=(0.3,0.3,0.3,1))
        btn_back.bind(on_press=self.back)
        self.lbl_summary = Label(text="Loading...", markup=True, halign='center')
        btn_exp = Button(text="SAVE\nEXCEL", size_hint_x=None, width=dp(80), background_color=COLOR_ACCENT, bold=True)
        btn_exp.bind(on_press=self.export)
        dash.add_widget(btn_back)
        dash.add_widget(self.lbl_summary)
        dash.add_widget(btn_exp)
        self.layout.add_widget(dash)
        
        controls = BoxLayout(size_hint_y=None, height=dp(50), spacing=10)
        self.search_input = TextInput(hint_text="Search product...", size_hint_x=0.7, multiline=False)
        self.search_input.bind(text=self.filter_list)
        
        self.btn_view = Button(text="View: List", size_hint_x=0.3, background_color=(0.2,0.2,0.2,1))
        self.btn_view.bind(on_press=self.toggle_view)

        controls.add_widget(self.search_input)
        controls.add_widget(self.btn_view)
        self.layout.add_widget(controls)

        self.table_header = BoxLayout(size_hint_y=None, height=dp(30), spacing=10)
        self.table_header.add_widget(Label(text="Product", size_hint_x=0.5, color=COLOR_ACCENT))
        self.table_header.add_widget(Label(text="Cost", size_hint_x=0.2, color=COLOR_ACCENT))
        self.table_header.add_widget(Label(text="Qty", size_hint_x=0.15, color=COLOR_ACCENT))
        self.table_header.add_widget(Label(text="Total", size_hint_x=0.25, color=COLOR_ACCENT))
        self.table_header.opacity = 0 
        self.layout.add_widget(self.table_header)

        self.scroll = ScrollView()
        self.grid = GridLayout(cols=1, spacing=5, size_hint_y=None)
        self.grid.bind(minimum_height=self.grid.setter('height'))
        self.scroll.add_widget(self.grid)
        self.layout.add_widget(self.scroll)
        self.add_widget(self.layout)

    def back(self, i): self.manager.current = 'home'
    def export(self, i):
        success, name = export_results_smart()
        if success: Popup(title="Success", content=Label(text=f"Saved:\n{name}"), size_hint=(0.7,0.4)).open()
        else: Popup(title="Error", content=Label(text=str(name)), size_hint=(0.8,0.4)).open()
    
    def toggle_view(self, instance):
        if self.view_mode == "list":
            self.view_mode = "table"
            self.btn_view.text = "View: Table"
            self.table_header.opacity = 1
            self.grid.cols = 1
        elif self.view_mode == "table":
            self.view_mode = "gallery"
            self.btn_view.text = "View: Gallery"
            self.table_header.opacity = 0
            self.grid.cols = 2
        else:
            self.view_mode = "list"
            self.btn_view.text = "View: List"
            self.table_header.opacity = 0
            self.grid.cols = 1
        self.load_data() 
        
    def filter_list(self, instance, value):
        self.load_data(filter_text=value)
        
    def load_data(self, filter_text=""):
        self.grid.clear_widgets()
        total_d = SESSION_STATE.get("total_investment", 0)
        count = len(SESSION_STATE["data"])
        self.lbl_summary.text = f"[b]{count} Items[/b]\nTotal: [color=00cc66]{int(total_d):,} DA[/color]"
        filter_text = filter_text.lower()
        
        for item in SESSION_STATE["data"]:
            if filter_text and filter_text not in item['name'].lower(): continue
            
            if self.view_mode == "list":
                self.grid.add_widget(InfoCard(item))
                self.grid.add_widget(Button(size_hint_y=None, height=1, background_color=(0.3,0.3,0.3,1)))
            elif self.view_mode == "table":
                self.grid.add_widget(TableRow(item))
            elif self.view_mode == "gallery":
                self.grid.add_widget(GalleryCard(item))

class SettingsScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=20, spacing=20)
        layout.add_widget(Label(text="SETTINGS", font_size='24sp', color=COLOR_ACCENT, size_hint=(1, 0.1)))
        self.inputs = {}
        fields = [("Exchange Rate", "exchange_rate"), ("Shipping (DZD/CBM)", "shipping_rate")]
        for lbl, key in fields:
            box = BoxLayout(orientation='vertical', size_hint=(1, 0.2))
            box.add_widget(Label(text=lbl, size_hint_y=0.4, halign='left', text_size=(Window.width-40, None)))
            inp = TextInput(text=str(GLOBAL_SETTINGS[key]), multiline=False, background_color=(0.2,0.2,0.2,1), foreground_color=(1,1,1,1))
            self.inputs[key] = inp
            box.add_widget(inp)
            layout.add_widget(box)
        btn_save = Button(text="SAVE", size_hint=(1, 0.15), background_color=COLOR_ACCENT, bold=True)
        btn_save.bind(on_press=self.save)
        layout.add_widget(btn_save)
        self.add_widget(layout)
    def save(self, instance):
        try:
            GLOBAL_SETTINGS["exchange_rate"] = float(self.inputs["exchange_rate"].text)
            GLOBAL_SETTINGS["shipping_rate"] = float(self.inputs["shipping_rate"].text)
            self.manager.current = 'home'
        except: pass

class ImportApp(App):
    def build(self):
        Window.clearcolor = COLOR_BG
        sm = ScreenManager()
        sm.add_widget(HomeScreen(name='home'))
        sm.add_widget(SettingsScreen(name='settings'))
        sm.add_widget(ResultsScreen(name='results'))
        Clock.schedule_once(lambda dt: request_android_permissions(), 1)
        return sm

if __name__ == '__main__':
    ImportApp().run()
